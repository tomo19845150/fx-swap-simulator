from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "7月の取り組み"

# A4横：297mm x 210mm
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 9  # A4
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3

# 列幅・行高さ設定（単位：文字幅 / ポイント）
col_widths = [1.5, 8, 8, 8, 1.5, 8, 8, 8, 1.5, 8, 8, 8, 1.5]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row_heights = [6, 20, 8, 10, 30, 10, 20, 10, 30, 10, 20, 10, 36, 8, 14, 6]
for i, h in enumerate(row_heights, 1):
    ws.row_dimensions[i].height = h

# ヘルパー
def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def cell(ws, row, col):
    return ws.cell(row=row, column=col)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color="000000", size=11, bold=False):
    return Font(name="メイリオ", color=color, size=size, bold=bold)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style="thin", color="CCCCCC")
thick = Side(style="medium", color="FFFFFF")
no_border = Border()

def card_border(top_color):
    top_side = Side(style="thick", color=top_color)
    return Border(top=top_side,
                  left=Side(style="thin", color="DDDDDD"),
                  right=Side(style="thin", color="DDDDDD"),
                  bottom=Side(style="thin", color="DDDDDD"))

# ===== 背景：全体を薄青 =====
for r in range(1, 17):
    for c in range(1, 14):
        ws.cell(r, c).fill = fill("EBF5FB")

# ===== ヘッダー (rows 2-5, cols 2-12) =====
merge(ws, 2, 2, 5, 12)
c = cell(ws, 2, 2)
c.value = "２０２６年 ７月の取り組み\n衛生管理・見える化プロジェクト"
c.fill = fill("1A5276")
c.font = Font(name="メイリオ", color="FFFFFF", size=24, bold=True)
c.alignment = align("center", "center", True)

# ===== カード共通関数 =====
def make_card(icon_row, label_row, main_row, sub_row,
              col_s, col_e,
              icon, label, main_text, sub_text,
              top_color, bg_color, label_color, main_color="1C1C1C"):
    # カード背景
    for r in range(icon_row, sub_row + 1):
        for c in range(col_s, col_e + 1):
            ws.cell(r, c).fill = fill(bg_color)

    # アイコン
    merge(ws, icon_row, col_s, icon_row, col_e)
    c = cell(ws, icon_row, col_s)
    c.value = icon
    c.font = Font(name="Segoe UI Emoji", size=18)
    c.alignment = align("left", "bottom", False)
    c.fill = fill(bg_color)

    # ラベル
    merge(ws, label_row, col_s, label_row, col_e)
    c = cell(ws, label_row, col_s)
    c.value = label
    c.font = Font(name="メイリオ", color=label_color, size=8, bold=True)
    c.alignment = align("left", "center", False)
    c.fill = fill(bg_color)

    # メインテキスト
    merge(ws, main_row, col_s, main_row, col_e)
    c = cell(ws, main_row, col_s)
    c.value = main_text
    c.font = Font(name="メイリオ", color=main_color, size=13, bold=True)
    c.alignment = align("left", "center", True)
    c.fill = fill(bg_color)

    # サブテキスト
    merge(ws, sub_row, col_s, sub_row, col_e)
    c = cell(ws, sub_row, col_s)
    c.value = sub_text
    c.font = Font(name="メイリオ", color="555555", size=8)
    c.alignment = align("left", "top", True)
    c.fill = fill(bg_color)

    # 上辺ボーダー（太線）
    for col in range(col_s, col_e + 1):
        ws.cell(icon_row, col).border = Border(
            top=Side(style="thick", color=top_color)
        )

# ===== カード① 殺菌ダスタ (cols 2-4) =====
make_card(7, 9, 10, 11, 2, 4,
          "🧴", "【取り組み①】衛生管理",
          "殺菌ダスタの運用を定着させる",
          "正しい手順での殺菌ダスタ使用を徹底し、全スタッフへの周知・習慣化を図る。",
          "2980B9", "FFFFFF", "1A5276")

# ===== カード② バイジング (cols 6-8) =====
make_card(7, 9, 10, 11, 6, 8,
          "👤", "【取り組み②】実施者管理",
          "バイジングシステムで実施者を明確化",
          "バイジングシステムへの入力を徹底し、誰が・いつ・何を行ったかを記録・管理する。",
          "2980B9", "FFFFFF", "1A5276")

# ===== カード③ 見える化 (cols 10-12) =====
make_card(7, 9, 10, 11, 10, 12,
          "✅", "【目標】作業の見える化",
          "実施状況をリアルタイムで把握",
          "チェックリスト・記録表を活用し、実施・未実施を誰でも一目で確認できる仕組みを構築。",
          "27AE60", "F0FFF4", "1E8449")

# ===== 課題カード (cols 2-8) =====
make_card(12, 13, 14, 14, 2, 8,
          "⚠️", "【現状の課題】",
          "清掃管理が未実施 ― 机上の空論で実際には行われていない",
          "ルール・手順は存在するにもかかわらず、現場では実施されていません。今月より「仕組みで動く」体制へ切り替えます。",
          "E74C3C", "FFF5F5", "C0392B")

# ===== ゴールカード (cols 10-12) =====
make_card(12, 13, 14, 14, 10, 12,
          "🎯", "【今月のゴール】",
          "衛生管理を確実に実施する",
          "記録・管理・確認の三本柱で衛生管理を現場に根付かせる。",
          "27AE60", "F0FFF4", "1E8449")

# ===== フッター (row 15, cols 2-12) =====
merge(ws, 15, 2, 15, 12)
c = cell(ws, 15, 2)
c.value = "▶ 全員で取り組む衛生管理 ― まず「見える化」から始めよう"
c.fill = fill("1A5276")
c.font = Font(name="メイリオ", color="FFFFFF", size=12, bold=True)
c.alignment = align("center", "center", False)

out = r"c:\Users\exper\マイドライブ\AIの作業場\html_tools\poster_july_hygiene.xlsx"
wb.save(out)
print("完成:", out)
